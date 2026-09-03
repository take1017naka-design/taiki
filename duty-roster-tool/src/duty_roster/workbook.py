"""勤務割当表(Excel)の読み取り。

実際の勤務割当表は次のようなレイアウトになっている::

    行5 |          | 日   | 1  | 2  | 3  | ... | 31 |
    行6 | 氏 名    | 曜日 | 土 | 日 | 月 | ... | 月 |
    行7 | 31.5     | 予   | 公 | 公 |    |     |    |   <- 上段（予）
    行8 | 担当A 太郎| 実   |    |    |    |     |    |   <- 下段（実、ここに氏名）
    行9 | 21.5     | 予   |―/公| 公 | 機 |     | 内視|
    行10| 担当B 花子| 実   | ME |    |カテ|     |    |

* 日付見出し行・氏名列・区分(予/実)列は自動検出する（設定で固定も可能）。
* 氏名は上下段のどちらに書かれていてもよい（2行を1ブロックとして扱う）。
* 上下段（予・実）は運用上区別しないため、既定では両方をまとめて
  「その日の勤務記号」として扱う（`priority.duty_source` で変更可）。
* 文字色（赤字＝本人希望の不在）とセル背景色（濃い黄色＝待機不可、
  薄い黄緑＝対象外）を読み取る。テーマ色＋tint も解決する。
"""

from __future__ import annotations

import calendar
import colorsys
import datetime as dt
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

from .config import Config, normalize_code, normalize_name


class ScheduleError(Exception):
    """勤務表の読み取りに失敗した。"""


# xlsx の theme インデックス → clrScheme の要素名（Excel は lt1/dk1 が入れ替わる）
THEME_ORDER = [
    "lt1", "dk1", "lt2", "dk2",
    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
    "hlink", "folHlink",
]


def load_theme_colors(path: Path) -> list[tuple[int, int, int]]:
    """xl/theme/theme1.xml から配色を読み、theme インデックス順の RGB を返す。"""
    try:
        with zipfile.ZipFile(path) as zf:
            name = next((n for n in zf.namelist() if n.startswith("xl/theme/")), None)
            if name is None:
                return []
            xml = zf.read(name).decode("utf-8", errors="replace")
    except (OSError, zipfile.BadZipFile, StopIteration):
        return []

    block = re.search(r"<a:clrScheme.*?</a:clrScheme>", xml, re.S)
    if not block:
        return []
    scheme: dict[str, tuple[int, int, int]] = {}
    for tag, body in re.findall(r"<a:(\w+)>(.*?)</a:\1>", block.group(0), re.S):
        m = re.search(r'(?:srgbClr val|lastClr)="([0-9A-Fa-f]{6})"', body)
        if m:
            v = m.group(1)
            scheme[tag] = (int(v[0:2], 16), int(v[2:4], 16), int(v[4:6], 16))
    return [scheme.get(k, (0, 0, 0)) for k in THEME_ORDER]


def apply_tint(rgb: tuple[int, int, int], tint: float) -> tuple[int, int, int]:
    """ECMA-376 の tint を HLS の明度に適用する。"""
    if not tint:
        return rgb
    r, g, b = (c / 255 for c in rgb)
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = l * (1 + tint) if tint < 0 else l * (1 - tint) + tint
    l = min(1.0, max(0.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return (round(r * 255), round(g * 255), round(b * 255))


def resolve_rgb(color: Any, theme: list[tuple[int, int, int]]) -> tuple[int, int, int] | None:
    """openpyxl の Color から (r, g, b) を取り出す。判定不能なら None。"""
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        raw = getattr(color, "rgb", None)
        if isinstance(raw, str) and len(raw) >= 6:
            v = raw[-6:]
            try:
                return int(v[0:2], 16), int(v[2:4], 16), int(v[4:6], 16)
            except ValueError:
                return None
        return None
    if ctype == "indexed":
        idx = getattr(color, "indexed", None)
        if isinstance(idx, int) and 0 <= idx < len(COLOR_INDEX):
            v = str(COLOR_INDEX[idx])[-6:]
            try:
                return int(v[0:2], 16), int(v[2:4], 16), int(v[4:6], 16)
            except ValueError:
                return None
        return None
    if ctype == "theme":
        idx = getattr(color, "theme", None)
        if isinstance(idx, int) and 0 <= idx < len(theme):
            return apply_tint(theme[idx], float(getattr(color, "tint", 0.0) or 0.0))
    return None


def is_red(rgb: tuple[int, int, int] | None, colors: dict[str, Any]) -> bool:
    if rgb is None:
        return False
    r, g, b = rgb
    return (
        r >= int(colors["red_min_r"])
        and g <= int(colors["red_max_g"])
        and b <= int(colors["red_max_b"])
    )


def classify_fill_rgb(rgb: tuple[int, int, int] | None, colors: dict[str, Any]) -> str:
    """セル背景を "yellow" / "other" / "none" に分類する。

    「濃い黄色」だけを yellow とする。薄い黄緑・グレー・淡色は other/none。
    """
    if rgb is None:
        return "none"
    r, g, b = rgb
    if (r, g, b) in ((255, 255, 255), (0, 0, 0)):
        return "none"
    h, s, v = colorsys.rgb_to_hsv(r / 255, g / 255, b / 255)
    hue = h * 360
    lo, hi = colors["yellow_hue"]
    if (
        lo <= hue <= hi
        and s >= float(colors["yellow_min_saturation"])
        and v >= float(colors["yellow_min_value"])
    ):
        return "yellow"
    return "other"


# ---------------------------------------------------------------------------
# データ構造
# ---------------------------------------------------------------------------
@dataclass
class Cell:
    """勤務表の1セル。"""

    text: str = ""
    red: bool = False
    yellow: bool = False
    row_kind: str = ""  # "予" / "実"

    @property
    def blank(self) -> bool:
        return self.text == ""


@dataclass
class DayCells:
    """ある人のある日のセル（上段＝予・下段＝実の両方）。"""

    cells: list[Cell] = field(default_factory=list)

    def texts(self, source: str = "both") -> list[str]:
        out = []
        for c in self.cells:
            if source != "both" and c.row_kind != source:
                continue
            if c.text:
                out.append(c.text)
        return out

    @property
    def yellow(self) -> bool:
        return any(c.yellow for c in self.cells)

    @property
    def red(self) -> bool:
        return any(c.red and c.text for c in self.cells)


@dataclass
class WorkSchedule:
    """勤務割当表1か月分。"""

    year: int
    month: int
    cells: dict[tuple[str, dt.date], DayCells]
    names_in_sheet: list[str]
    source: Path | None = None
    warnings: list[str] = field(default_factory=list)

    def day_cells(self, name: str, day: dt.date) -> DayCells:
        return self.cells.get((name, day), DayCells())

    def has_day(self, name: str, day: dt.date) -> bool:
        return (name, day) in self.cells


# ---------------------------------------------------------------------------
# 読み取り
# ---------------------------------------------------------------------------
def _cell_day_number(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.day
    if isinstance(value, (int, float)) and float(value).is_integer():
        n = int(value)
        return n if 1 <= n <= 31 else None
    if isinstance(value, str):
        text = value.strip().rstrip("日")
        if text.isdigit():
            n = int(text)
            return n if 1 <= n <= 31 else None
    return None


def _detect_header_row(ws: Any, max_rows: int) -> tuple[int, list[tuple[int, int]]]:
    best: tuple[int, list[tuple[int, int]]] | None = None
    limit_rows = min(max_rows, ws.max_row or max_rows)
    limit_cols = ws.max_column or 40
    for row in range(1, limit_rows + 1):
        found = [
            (col, day)
            for col in range(1, limit_cols + 1)
            if (day := _cell_day_number(ws.cell(row=row, column=col).value)) is not None
        ]
        if len(found) >= 20 and (best is None or len(found) > len(best[1])):
            best = (row, found)
    if best is None:
        raise ScheduleError(
            "日付の見出し行（1〜31が並ぶ行）を検出できませんでした。"
            " excel.header_row で行番号を指定してください。"
        )
    return best


def _detect_kind_column(
    ws: Any, header_row: int, last_row: int, first_day_col: int, labels: set[str]
) -> int | None:
    best_col, best_hits = None, 0
    for col in range(1, first_day_col):
        hits = sum(
            1
            for row in range(header_row + 1, last_row + 1)
            if normalize_code(ws.cell(row=row, column=col).value) in labels
        )
        if hits > best_hits:
            best_col, best_hits = col, hits
    return best_col


def _match_member(text: str, members: list[str]) -> str | None:
    """セルの氏名表記から対象者を特定する（姓の前方一致）。"""
    key = normalize_name(text)
    if not key:
        return None
    for name in members:
        n = normalize_name(name)
        if key == n or key.startswith(n):
            return name
    return None


def _detect_name_column(
    ws: Any, header_row: int, last_row: int, first_day_col: int, members: list[str]
) -> int | None:
    best_col, best_hits = None, 0
    for col in range(1, first_day_col):
        hits = sum(
            1
            for row in range(header_row + 1, last_row + 1)
            if _match_member(str(ws.cell(row=row, column=col).value or ""), members)
        )
        if hits > best_hits:
            best_col, best_hits = col, hits
    return best_col


def detect_year_month(path: str | Path) -> tuple[int, int] | None:
    """表題（例: 「２０２６年　８月勤務割当表」）から年月を推定する。"""
    wb = load_workbook(Path(path), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=6):
            for cell in row:
                text = normalize_code(cell.value)
                m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", text)
                if m:
                    return int(m.group(1)), int(m.group(2))
    finally:
        wb.close()
    return None


def read_schedule(path: str | Path, year: int, month: int, cfg: Config) -> WorkSchedule:
    """勤務割当表 Excel を読み取る。"""
    path = Path(path)
    if not path.exists():
        raise ScheduleError(f"勤務表が見つかりません: {path}")

    opts = cfg.excel
    theme = load_theme_colors(path)
    wb = load_workbook(path)  # 書式が必要なので data_only=False で開く
    if opts.get("sheet"):
        if opts["sheet"] not in wb.sheetnames:
            raise ScheduleError(
                f"シート '{opts['sheet']}' がありません。存在するシート: {wb.sheetnames}"
            )
        ws = wb[opts["sheet"]]
    else:
        ws = wb[wb.sheetnames[0]]

    if opts.get("header_row"):
        header_row = int(opts["header_row"])
        day_cols = [
            (col, day)
            for col in range(1, (ws.max_column or 60) + 1)
            if (day := _cell_day_number(ws.cell(row=header_row, column=col).value)) is not None
        ]
        if not day_cols:
            raise ScheduleError(f"指定された見出し行 {header_row} に日付が見つかりません。")
    else:
        header_row, day_cols = _detect_header_row(ws, int(opts["max_scan_rows"]))

    first_day_col = day_cols[0][0]
    last_row = ws.max_row or header_row
    members = cfg.member_names

    plan_label = normalize_code(opts["plan_label"])
    actual_label = normalize_code(opts["actual_label"])
    kind_col = opts.get("kind_column") or _detect_kind_column(
        ws, header_row, last_row, first_day_col, {plan_label, actual_label}
    )
    if not kind_col:
        raise ScheduleError(
            f"区分列（'{opts['plan_label']}' / '{opts['actual_label']}'）を検出できませんでした。"
            " excel.kind_column を指定してください。"
        )
    name_col = opts.get("name_column") or _detect_name_column(
        ws, header_row, last_row, first_day_col, members
    )
    if not name_col:
        raise ScheduleError(
            "氏名列を検出できませんでした。設定の members に書いた姓が勤務表の表記と"
            " 一致しているか確認するか、excel.name_column を指定してください。"
        )

    # 日付列 → 実日付。月末を跨いで翌月に続く場合にも対応する。
    days_in_month = calendar.monthrange(year, month)[1]
    col_to_date: dict[int, dt.date] = {}
    rolled = False
    prev_day = 0
    for col, day in day_cols:
        if day < prev_day:
            rolled = True
        prev_day = day
        if rolled or day > days_in_month:
            ny, nm = (year + 1, 1) if month == 12 else (year, month + 1)
            try:
                col_to_date[col] = dt.date(ny, nm, day)
            except ValueError:
                continue
        else:
            col_to_date[col] = dt.date(year, month, day)

    colors = cfg.colors
    aliases = cfg.code_aliases
    cells: dict[tuple[str, dt.date], DayCells] = {}
    names_in_sheet: list[str] = []
    warnings: list[str] = []

    # 予/実 の2行を1ブロックとして扱う。氏名は上下どちらの行にあってもよい。
    blocks: list[tuple[str | None, list[tuple[int, str]]]] = []
    current: list[tuple[int, str]] = []
    current_name: str | None = None
    for row in range(header_row + 1, last_row + 1):
        kind = normalize_code(ws.cell(row=row, column=kind_col).value)
        if kind not in (plan_label, actual_label):
            continue
        if kind == plan_label and current:
            blocks.append((current_name, current))
            current, current_name = [], None
        current.append((row, kind))
        matched = _match_member(str(ws.cell(row=row, column=name_col).value or ""), members)
        if matched:
            current_name = matched
    if current:
        blocks.append((current_name, current))

    for name, rows in blocks:
        if not name:
            continue
        if name not in names_in_sheet:
            names_in_sheet.append(name)
        for row, kind in rows:
            row_kind = "予" if kind == plan_label else "実"
            for col, day in col_to_date.items():
                cell = ws.cell(row=row, column=col)
                text = normalize_code(cell.value)
                text = aliases.get(text, text)
                font_rgb = resolve_rgb(getattr(cell.font, "color", None), theme)
                fill = getattr(cell, "fill", None)
                fill_rgb = (
                    resolve_rgb(fill.start_color, theme)
                    if fill is not None and fill.fill_type == "solid"
                    else None
                )
                info = Cell(
                    text=text,
                    red=is_red(font_rgb, colors),
                    yellow=classify_fill_rgb(fill_rgb, colors) == "yellow",
                    row_kind=row_kind,
                )
                cells.setdefault((name, day), DayCells()).cells.append(info)

    missing = [n for n in members if n not in names_in_sheet]
    if missing:
        warnings.append(
            f"勤務表に見つからなかった対象者: {', '.join(missing)}"
            "（設定の氏名表記を確認してください。全日勤務可能として扱います）"
        )
    if not any(d.month != month for d in col_to_date.values()):
        warnings.append(
            "翌月1日の列がないため、月末日の「翌日の勤務内容」は空白として扱います。"
        )

    return WorkSchedule(
        year=year,
        month=month,
        cells=cells,
        names_in_sheet=names_in_sheet,
        source=path,
        warnings=warnings,
    )


def list_member_names(path: str | Path, cfg: Config | None = None) -> list[str]:
    """勤務表から氏名らしき文字列を抽出する（init-config 用）。"""
    wb = load_workbook(Path(path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, day_cols = _detect_header_row(ws, 40)
    first_day_col = day_cols[0][0]
    names: list[str] = []
    for row in range(header_row + 1, (ws.max_row or header_row) + 1):
        for col in range(1, first_day_col):
            text = normalize_code(ws.cell(row=row, column=col).value)
            # 姓＋空白＋名 の形（数値や記号は除外）
            if re.fullmatch(r"[^\W\d_]{2,4} ?[^\W\d_]{1,4}", text) and text not in names:
                names.append(text)
    return names
