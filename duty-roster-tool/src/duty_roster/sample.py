"""動作確認用のダミー勤務割当表を生成する。

実データを使わずに読み取り〜割り当てまでを試すためのもの。
`config/roster.example.yaml` の氏名（担当A〜担当H）に合わせてある。
"""

from __future__ import annotations

import calendar
import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

SAMPLE_MEMBERS = ["担当A", "担当B", "担当C", "担当D", "担当E", "担当F", "担当G", "担当H"]
DUTIES = ["ME", "OHP", "内視", "機", "OP", "アーム", "カテ", "ABL", "会議", "O", ""]
WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]
RED = Font(color="FFFF0000")
YELLOW = PatternFill("solid", fgColor="FFFFFF00")


def build_sample(path: str | Path, year: int, month: int, seed: int = 20260801) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rng = random.Random(seed)
    days = calendar.monthrange(year, month)[1]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=2, column=8, value=f"臨床工学科　{year}年　{month}月勤務割当表（サンプル）")

    ws.cell(row=5, column=2, value="日")
    ws.cell(row=6, column=1, value="氏 名")
    ws.cell(row=6, column=2, value="曜日")
    for d in range(1, days + 1):
        col = 2 + d
        ws.cell(row=5, column=col, value=d).alignment = Alignment(horizontal="center")
        wd = dt.date(year, month, d).weekday()
        ws.cell(row=6, column=col, value=WEEKDAY_JP[wd]).alignment = Alignment(horizontal="center")

    row = 7
    for index, name in enumerate(SAMPLE_MEMBERS):
        ws.cell(row=row, column=2, value="予")
        ws.cell(row=row + 1, column=2, value="実")
        ws.cell(row=row + 1, column=1, value=f"{name} 太郎")
        for d in range(1, days + 1):
            col = 2 + d
            date = dt.date(year, month, d)
            wd = date.weekday()
            plan_cell = ws.cell(row=row, column=col)
            actual_cell = ws.cell(row=row + 1, column=col)

            if wd == 6:  # 日曜は全員 公（黒字＝出勤可能扱い）
                plan_cell.value = "公"
                continue
            roll = rng.random()
            if roll < 0.10:
                plan_cell.value = rng.choice(["有", "夏", "リ", "有/公"])
            elif roll < 0.18:
                plan_cell.value = "公"
                plan_cell.font = RED  # 本人希望の不在
            else:
                plan_cell.value = rng.choice(DUTIES)
                if rng.random() < 0.25:
                    actual_cell.value = rng.choice(DUTIES)
            if rng.random() < 0.02:
                actual_cell.fill = YELLOW  # 本人希望で待機不可
        row += 2

    wb.save(path)
    return path
