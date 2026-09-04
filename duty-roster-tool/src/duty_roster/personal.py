"""個人別の待機・予備表。

1人1シートで、その人が担当する日だけを書き出す。他の人の氏名は出さない。

* 待機の日 … 赤字で「待機」
* 予備の日 … 黒字で「予備」
* 担当不可の日（勤務表が赤字・黄色セル）… セルを黄色で塗る
* 翌日が手術室業務の担当日 … セルを緑色で塗る
* それ以外 … 空欄

書式・印刷設定は待機表と同じ（日曜始まり、A4横1ページ）。
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import Config
from .rules import RuleEngine
from .writer import BOX, CENTER, THIN, _day_fill, _day_font_color, setup_print

DUTY_LABEL = "待機"
BACKUP_LABEL = "予備"


def write_personal_rosters(
    path: str | Path,
    cfg: Config,
    engine: RuleEngine,
    primary: dict[dt.date, str],
    backup: dict[dt.date, str] | None = None,
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    backup = backup or {}

    wb = Workbook()
    wb.remove(wb.active)
    for name in engine.members:
        _sheet_for(wb, cfg, engine, name, primary, backup)
    wb.save(path)
    return path


def _sheet_for(
    wb: Workbook,
    cfg: Config,
    engine: RuleEngine,
    name: str,
    primary: dict[dt.date, str],
    backup: dict[dt.date, str],
) -> None:
    out = cfg.output
    ws = wb.create_sheet(name)

    duty_color = out.get("personal_duty_color", "FFFF0000")
    backup_color = out.get("personal_backup_color", "FF000000")
    unavailable_fill = out.get("personal_unavailable_fill", "FFFFF2CC")
    operating_room_fill = out.get("personal_operating_room_fill", "FFC6EFCE")

    title = f"{engine.year}年{engine.month}月　{name}　待機・予備"
    ws.cell(row=1, column=1, value=title).font = Font(
        size=int(out.get("title_font_size", 18)), bold=True
    )
    ws.row_dimensions[1].height = int(out.get("title_font_size", 18)) * 1.6

    header = ["日", "月", "火", "水", "木", "金", "土"]
    for i, label in enumerate(header, start=1):
        cell = ws.cell(row=3, column=i, value=label)
        cell.alignment = CENTER
        cell.border = BOX
        color = (
            out["sunday_font_color"]
            if i == 1
            else out["saturday_font_color"] if i == 7 else out["weekday_font_color"]
        )
        cell.font = Font(bold=True, color=color, size=int(out.get("weekday_font_size", 14)))
        cell.fill = PatternFill("solid", fgColor="FFF2F2F2")
        ws.column_dimensions[get_column_letter(i)].width = float(out["calendar_column_width"])
    ws.row_dimensions[3].height = int(out.get("weekday_font_size", 14)) * 1.6

    def column_of(day: dt.date) -> int:
        return (day.weekday() + 1) % 7 + 1

    row = 4
    prev_col = None
    duty_days: list[dt.date] = []
    backup_days: list[dt.date] = []
    for day in engine.days:
        col = column_of(day)
        if prev_col is not None and col <= prev_col:
            row += 2
        prev_col = col

        date_cell = ws.cell(row=row, column=col, value=day.day)
        date_cell.font = Font(
            bold=True,
            color=_day_font_color(engine, day, out),
            size=int(out.get("date_font_size", 14)),
        )
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        date_cell.border = Border(left=THIN, right=THIN, top=THIN)

        if primary.get(day) == name:
            label, color, bold = DUTY_LABEL, duty_color, True
            duty_days.append(day)
        elif backup.get(day) == name:
            label, color, bold = BACKUP_LABEL, backup_color, False
            backup_days.append(day)
        else:
            label, color, bold = None, None, False

        role_cell = ws.cell(row=row + 1, column=col, value=label)
        role_cell.alignment = CENTER
        role_cell.border = Border(left=THIN, right=THIN, bottom=THIN)
        if label:
            role_cell.font = Font(
                color=color, bold=bold, size=int(out.get("name_font_size", 18))
            )

        # 担当日で翌日が手術室業務なら緑、本人希望の担当不可なら黄色で塗る
        if label and operating_room_fill and engine.next_day_is_operating_room(name, day):
            special = operating_room_fill
        elif not label and unavailable_fill and engine.is_personally_unavailable(name, day):
            special = unavailable_fill
        else:
            special = None
        if special:
            date_cell.fill = PatternFill("solid", fgColor=special)
            role_cell.fill = PatternFill("solid", fgColor=special)
            continue
        fill = _day_fill(engine, day, out)
        if fill is not None:
            date_cell.fill = fill
            role_cell.fill = PatternFill("solid", fgColor=fill.start_color.rgb)

    date_height = float(out["calendar_date_row_height"])
    name_height = float(out["calendar_name_row_height"])
    for r in range(4, row + 2, 2):
        ws.row_dimensions[r].height = date_height
        ws.row_dimensions[r + 1].height = name_height

    ws.cell(
        row=row + 3,
        column=1,
        value=(
            f"※ 赤字＝{DUTY_LABEL} / 黒字＝{BACKUP_LABEL}"
            " / 緑のセル＝翌日が手術室業務 / 黄色のセル＝担当不可（希望）"
        ),
    ).font = Font(size=9, color="FF808080")

    _counts(ws, cfg, duty_days, backup_days, duty_color, backup_color)

    last_col = int(cfg.output.get("counts_column", 9)) + 1
    setup_print(
        ws, cfg, area=f"A1:{get_column_letter(last_col)}{row + 3}", fit_height=1
    )


def _counts(ws, cfg: Config, duty_days, backup_days, duty_color, backup_color) -> None:
    """カレンダーの右横に、その人の待機・予備・合計の日数を出す。"""
    col = int(cfg.output.get("counts_column", 9))
    row = int(cfg.output.get("counts_row", 3))
    size = int(cfg.output.get("counts_font_size", 12))

    ws.column_dimensions[get_column_letter(col - 1)].width = 2.5
    ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions[get_column_letter(col + 1)].width = 8

    for offset, label in enumerate(("区分", "日数")):
        cell = ws.cell(row=row, column=col + offset, value=label)
        cell.font = Font(bold=True, size=size)
        cell.alignment = CENTER
        cell.border = BOX
        cell.fill = PatternFill("solid", fgColor="FFF2F2F2")

    rows = [
        ("合計", len(duty_days) + len(backup_days), "FF000000", True),
        (DUTY_LABEL, len(duty_days), duty_color, True),
        (BACKUP_LABEL, len(backup_days), backup_color, False),
    ]
    for index, (label, count, color, bold) in enumerate(rows, start=1):
        label_cell = ws.cell(row=row + index, column=col, value=label)
        label_cell.font = Font(size=size, color=color, bold=bold)
        label_cell.border = BOX
        label_cell.alignment = Alignment(horizontal="left", vertical="center")

        count_cell = ws.cell(row=row + index, column=col + 1, value=count)
        count_cell.font = Font(size=size, color=color, bold=bold)
        count_cell.border = BOX
        count_cell.alignment = CENTER

    start = row + len(rows) + 2
    ws.cell(row=start, column=col, value="担当日").font = Font(bold=True, size=size)
    for index, (label, days, color) in enumerate(
        ((DUTY_LABEL, duty_days, duty_color), (BACKUP_LABEL, backup_days, backup_color)),
        start=1,
    ):
        cell = ws.cell(
            row=start + index,
            column=col,
            value=f"{label} " + " ".join(str(d.day) for d in days),
        )
        cell.font = Font(size=max(9, size - 2), color=color)
        cell.alignment = Alignment(vertical="top", wrap_text=False)
