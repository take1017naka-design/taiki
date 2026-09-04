"""月をまたいだ実績の記録。

日曜・祝日の予備待機を年間である程度均等にするため、月ごとの実績を
JSON に残しておき、次の月を組むときに参照する。

同じ月を作り直したときは、その月の記録を置き換える（二重に数えない）。
氏名を含むファイルなので、リポジトリには入れない（.gitignore 済み）。
"""

from __future__ import annotations

import datetime as dt
import json
from pathlib import Path
from typing import Any

CURRENT_VERSION = 1


class History:
    def __init__(self, path: Path, data: dict[str, Any] | None = None):
        self.path = path
        self.data = data or {"version": CURRENT_VERSION, "holiday_backup": {}}

    # -- 読み書き ----------------------------------------------------------
    @classmethod
    def load(cls, path: str | Path) -> "History":
        path = Path(path).expanduser()
        if not path.exists():
            return cls(path)
        try:
            with path.open(encoding="utf-8") as fh:
                data = json.load(fh)
        except (OSError, json.JSONDecodeError):
            return cls(path)
        data.setdefault("version", CURRENT_VERSION)
        data.setdefault("holiday_backup", {})
        return cls(path, data)

    def save(self) -> Path:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self.path.open("w", encoding="utf-8") as fh:
            json.dump(self.data, fh, ensure_ascii=False, indent=2, sort_keys=True)
            fh.write("\n")
        return self.path

    # -- 日曜・祝日の予備 --------------------------------------------------
    @staticmethod
    def _key(year: int, month: int) -> str:
        return f"{year:04d}-{month:02d}"

    def holiday_backup_totals(self, exclude_month: tuple[int, int] | None = None) -> dict[str, int]:
        """これまでの日曜・祝日の予備回数（指定した月を除く）。"""
        skip = self._key(*exclude_month) if exclude_month else None
        totals: dict[str, int] = {}
        for key, counts in self.data["holiday_backup"].items():
            if key == skip:
                continue
            for name, count in counts.items():
                totals[name] = totals.get(name, 0) + int(count)
        return totals

    def record_holiday_backup(
        self, year: int, month: int, assignment: dict[dt.date, str], is_target
    ) -> dict[str, int]:
        """その月の日曜・祝日の予備を記録する（同じ月は置き換える）。"""
        counts: dict[str, int] = {}
        for day, name in assignment.items():
            if is_target(day):
                counts[name] = counts.get(name, 0) + 1
        self.data["holiday_backup"][self._key(year, month)] = counts
        return counts

    def import_backup_roster(self, path) -> tuple[int, int, dict[str, int]]:
        """作成済みの予備待機表（.xlsx）から、その月の実績を取り込む。

        実績ファイル（history.json）を持ち回れないときに、
        すでに配ってある予備待機表から通算回数を組み直すために使う。
        """
        from openpyxl import load_workbook

        from .config import normalize_name
        from .holidays_jp import holidays_in_month

        wb = load_workbook(Path(path), data_only=True)
        if "一覧" not in wb.sheetnames:
            raise ValueError(
                f"{Path(path).name} に「一覧」シートがありません。"
                "このツールで作った予備待機表を渡してください。"
            )
        ws = wb["一覧"]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            date_col = header.index("日付") + 1
            name_col = header.index("担当") + 1
        except ValueError as exc:
            raise ValueError(
                f"{Path(path).name} の「一覧」シートに「日付」「担当」の列がありません。"
            ) from exc

        assignment: dict[dt.date, str] = {}
        for row in range(2, ws.max_row + 1):
            raw = ws.cell(row=row, column=date_col).value
            name = ws.cell(row=row, column=name_col).value
            if raw is None or not name:
                continue
            if isinstance(raw, dt.datetime):
                day = raw.date()
            elif isinstance(raw, dt.date):
                day = raw
            else:
                day = dt.datetime.strptime(str(raw).strip(), "%Y/%m/%d").date()
            assignment[day] = normalize_name(str(name))
        if not assignment:
            raise ValueError(f"{Path(path).name} から担当を読み取れませんでした。")

        first = min(assignment)
        year, month = first.year, first.month
        holidays = set(holidays_in_month(year, month))

        def is_red(day: dt.date) -> bool:
            return day.weekday() == 6 or day in holidays

        counts = self.record_holiday_backup(year, month, assignment, is_red)
        return year, month, counts

    def recorded_months(self) -> list[str]:
        return sorted(self.data["holiday_backup"])
