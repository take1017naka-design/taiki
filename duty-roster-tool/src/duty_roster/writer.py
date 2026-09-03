"""待機表 Excel の書き出し。"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from .config import Config
from .rules import SAT, SUN, WEEKDAY_JP, RuleEngine
from .solver import Solution

def setup_print(
    ws,
    cfg: Config,
    *,
    area: str | None = None,
    fit_height: int | None = 1,
    repeat_rows: str | None = None,
    center_vertically: bool | None = None,
) -> None:
    """そのまま印刷できるようにページ設定を入れる（既定 A4 横・幅を1ページに収める）。"""
    out = cfg.output
    ws.page_setup.paperSize = int(out.get("paper_size", 9))  # 9 = A4
    ws.page_setup.orientation = (
        "landscape" if out.get("paper_landscape", True) else "portrait"
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0 if fit_height is None else fit_height
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    margin = float(out.get("page_margin_inch", 0.4))
    ws.page_margins = PageMargins(
        left=margin, right=margin, top=margin, bottom=margin, header=0.2, footer=0.2
    )
    # 用紙の中央に配置する。複数ページになるシートは横方向だけ中央にする。
    ws.print_options.horizontalCentered = bool(out.get("center_horizontally", True))
    if center_vertically is None:
        center_vertically = fit_height == 1
    ws.print_options.verticalCentered = bool(
        center_vertically and out.get("center_vertically", True)
    )
    if area:
        ws.print_area = area
    if repeat_rows:
        ws.print_title_rows = repeat_rows


THIN = Side(style="thin", color="FF999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def _day_font_color(engine: RuleEngine, day: dt.date, out: dict[str, str]) -> str:
    if engine.is_red_day(day):
        return out["sunday_font_color"]
    if day.weekday() == SAT:
        return out["saturday_font_color"]
    return out["weekday_font_color"]


def _day_fill(engine: RuleEngine, day: dt.date, out: dict[str, str]) -> PatternFill | None:
    """日曜・祝日は同じ色、土曜は薄い青。"""
    if engine.is_red_day(day):
        color = out.get("holiday_fill")
    elif day.weekday() == SAT:
        color = out.get("saturday_fill")
    else:
        color = out.get("weekday_fill")
    return PatternFill("solid", fgColor=color) if color else None


def write_roster(
    path: str | Path,
    cfg: Config,
    engine: RuleEngine,
    solution: Solution,
    warnings: list[str] | None = None,
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    _sheet_calendar(wb, cfg, engine, solution)
    _sheet_list(wb, cfg, engine, solution)
    _sheet_summary(wb, cfg, engine, solution)
    _sheet_availability(wb, cfg, engine)
    _sheet_notes(wb, cfg, solution, warnings or [])

    wb.save(path)
    return path


def _sheet_calendar(wb: Workbook, cfg: Config, engine: RuleEngine, solution: Solution) -> None:
    ws = wb.active
    ws.title = "待機表"
    out = cfg.output

    ws.cell(row=1, column=1, value=f"{engine.year}年{engine.month}月　カテ待機表").font = Font(
        size=14, bold=True
    )
    header = ["日", "月", "火", "水", "木", "金", "土"]
    for i, label in enumerate(header, start=1):
        cell = ws.cell(row=3, column=i, value=label)
        cell.alignment = CENTER
        cell.border = BOX
        color = (
            out["sunday_font_color"]
            if i == 1
            else out["saturday_font_color"] if i == 7 else out["weekday_font_color"]
        )
        cell.font = Font(bold=True, color=color)
        cell.fill = PatternFill("solid", fgColor="FFF2F2F2")
        ws.column_dimensions[get_column_letter(i)].width = float(out["calendar_column_width"])

    # 日曜始まりの列位置（Python の weekday は月=0 なので +1 して 7 で割る）
    def column_of(day: dt.date) -> int:
        return (day.weekday() + 1) % 7 + 1

    row = 4
    prev_col = None
    for day in engine.days:
        col = column_of(day)
        if prev_col is not None and col <= prev_col:
            row += 2
        prev_col = col
        date_cell = ws.cell(row=row, column=col, value=day.day)
        date_cell.font = Font(bold=True, color=_day_font_color(engine, day, out))
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        date_cell.border = Border(left=THIN, right=THIN, top=THIN)

        name_cell = ws.cell(row=row + 1, column=col, value=solution.assignment[day])
        name_cell.alignment = CENTER
        name_cell.border = Border(left=THIN, right=THIN, bottom=THIN)
        fill = _day_fill(engine, day, out)
        if fill is not None:
            date_cell.fill = fill
            name_cell.fill = PatternFill("solid", fgColor=fill.start_color.rgb)

    date_height = float(out["calendar_date_row_height"])
    name_height = float(out["calendar_name_row_height"])
    for r in range(4, row + 2, 2):
        ws.row_dimensions[r].height = date_height
        ws.row_dimensions[r + 1].height = name_height

    ws.cell(row=row + 3, column=1, value="※ 日付の色: 平日=黒 / 土曜=青 / 日曜・祝日=赤").font = Font(
        size=9, color="FF808080"
    )

    _counts_block(ws, cfg, engine, solution)

    last_col = int(cfg.output.get("counts_column", 9)) + 1
    setup_print(
        ws,
        cfg,
        area=f"A1:{get_column_letter(last_col)}{row + 3}",
        fit_height=1,
    )


def _counts_block(ws, cfg: Config, engine: RuleEngine, solution: Solution) -> None:
    """カレンダーの右横に、担当ごとの待機日数を出す。"""
    col = int(cfg.output.get("counts_column", 9))
    row = int(cfg.output.get("counts_row", 3))
    quota = cfg.quota(engine.days_in_month)

    ws.column_dimensions[get_column_letter(col - 1)].width = 2.5
    ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions[get_column_letter(col + 1)].width = 7

    for offset, label in enumerate(("担当", "日数")):
        cell = ws.cell(row=row, column=col + offset, value=label)
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = BOX
        cell.fill = PatternFill("solid", fgColor="FFF2F2F2")

    for index, name in enumerate(engine.members, start=1):
        count = solution.counts.get(name, 0)
        target = quota.get(name, 0)
        name_cell = ws.cell(row=row + index, column=col, value=name)
        name_cell.border = BOX
        name_cell.alignment = Alignment(horizontal="left", vertical="center")

        count_cell = ws.cell(row=row + index, column=col + 1, value=count)
        count_cell.border = BOX
        count_cell.alignment = CENTER
        if count != target:
            # 目標と食い違っていたら赤字で目立たせる（通常は起こらない）
            count_cell.font = Font(color="FFFF0000", bold=True)

    total_row = row + len(engine.members) + 1
    total = ws.cell(row=total_row, column=col, value="合計")
    total.font = Font(bold=True)
    total.border = BOX
    total.alignment = Alignment(horizontal="left", vertical="center")
    total_value = ws.cell(row=total_row, column=col + 1, value=sum(solution.counts.values()))
    total_value.font = Font(bold=True)
    total_value.border = BOX
    total_value.alignment = CENTER


def _sheet_list(wb: Workbook, cfg: Config, engine: RuleEngine, solution: Solution) -> None:
    ws = wb.create_sheet("一覧")
    headers = ["日付", "曜日", "待機者", "採用した優先順位", "翌日の勤務", "同日の他候補"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2F2F2")
        c.border = BOX
    widths = [12, 6, 12, 20, 22, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r, day in enumerate(engine.days, start=2):
        name = solution.assignment[day]
        others = [n for n in engine.candidates(day) if n != name]
        values = [
            day.strftime("%Y/%m/%d"),
            WEEKDAY_JP[day.weekday()],
            name,
            engine.tier_label(name, day),
            engine.next_duty_text(name, day),
            "・".join(others),
        ]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = BOX
            if i == 2:
                cell.alignment = CENTER
    ws.freeze_panes = "A2"
    setup_print(ws, cfg, fit_height=None, repeat_rows="1:1")


def _sheet_summary(wb: Workbook, cfg: Config, engine: RuleEngine, solution: Solution) -> None:
    ws = wb.create_sheet("集計")
    quota = cfg.quota(engine.days_in_month)
    headers = ["氏名", "目標回数", "実績回数", "差", "土日祝", "連日", "待機日"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2F2F2")
        c.border = BOX
    for i, w in enumerate([12, 10, 10, 8, 8, 8, 46], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    days_of = {n: [] for n in engine.members}
    for day in engine.days:
        days_of[solution.assignment[day]].append(day)

    for r, name in enumerate(engine.members, start=2):
        mine = days_of[name]
        holiday_count = sum(1 for d in mine if engine.is_red_day(d) or d.weekday() == SAT)
        consec = sum(1 for d in mine if (d + dt.timedelta(days=1)) in mine)
        target = quota.get(name, 0)
        values = [
            name,
            target,
            len(mine),
            len(mine) - target,
            holiday_count,
            consec,
            " ".join(f"{d.day}({WEEKDAY_JP[d.weekday()]})" for d in mine),
        ]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = BOX
            if i == 4 and v != 0:
                cell.font = Font(color="FFFF0000", bold=True)

    setup_print(ws, cfg, fit_height=1, repeat_rows="1:1")


def _sheet_availability(wb: Workbook, cfg: Config, engine: RuleEngine) -> None:
    """読み取り結果の目視確認用。○=待機可 / 記号=不可理由。"""
    ws = wb.create_sheet("可否一覧")
    ws.cell(row=1, column=1, value="待機可否（○=可、△=条件付きで可、×=不可・下に理由）").font = Font(bold=True)
    ws.cell(row=2, column=1, value="氏名").font = Font(bold=True)
    for i, day in enumerate(engine.days, start=2):
        c = ws.cell(row=2, column=i, value=day.day)
        c.font = Font(bold=True)
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = 4.5
    ws.column_dimensions["A"].width = 12

    for r, name in enumerate(engine.members, start=3):
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        for i, day in enumerate(engine.days, start=2):
            elig = engine.eligibility(name, day)
            cell = ws.cell(row=r, column=i, value=engine.availability_mark(name, day))
            cell.alignment = CENTER
            if not elig.ok:
                cell.font = Font(color="FF999999")
            elif elig.conditional:
                cell.font = Font(color="FFC55A11", bold=True)
            if engine.is_yellow(name, day):
                cell.fill = PatternFill("solid", fgColor="FFFFFF00")
    ws.freeze_panes = "B3"
    setup_print(ws, cfg, fit_height=None, repeat_rows="2:2")

    start = len(engine.members) + 5
    ws.cell(row=start, column=1, value="不可・条件付きの理由").font = Font(bold=True)
    r = start + 1
    for name in engine.members:
        for day in engine.days:
            elig = engine.eligibility(name, day)
            if elig.ok and not elig.conditional:
                continue
            ws.cell(row=r, column=1, value=f"{day:%m/%d}")
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value="△" if elig.conditional else "×")
            ws.cell(row=r, column=4, value=elig.note if elig.conditional else elig.reason)
            r += 1


def _sheet_notes(wb: Workbook, cfg: Config, solution: Solution, warnings: list[str]) -> None:
    ws = wb.create_sheet("確認事項")
    ws.column_dimensions["A"].width = 90
    row = 1
    for title, items in (
        ("ルール違反・要確認", solution.violations),
        ("メモ", solution.notes),
        ("読み取りの注意", warnings),
    ):
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1
        if not items:
            ws.cell(row=row, column=1, value="（なし）")
            row += 2
            continue
        for item in items:
            ws.cell(row=row, column=1, value=f"・{item}")
            row += 1
        row += 1

    setup_print(ws, cfg, fit_height=None)
