import datetime as dt

import pytest
from openpyxl import load_workbook

from duty_roster.backup import solve_backup
from duty_roster.config import load_config
from duty_roster.personal import write_personal_rosters
from duty_roster.rules import RuleEngine
from duty_roster.sample import build_sample
from duty_roster.solver import solve
from duty_roster.workbook import read_schedule

CFG = load_config("config/roster.example.yaml")
YEAR, MONTH = 2026, 8


@pytest.fixture(scope="module")
def personal(tmp_path_factory):
    work = tmp_path_factory.mktemp("personal")
    engine = RuleEngine(
        CFG, read_schedule(build_sample(work / "s.xlsx", YEAR, MONTH), YEAR, MONTH, CFG),
        YEAR, MONTH,
    )
    primary = solve(CFG, engine)
    backup = solve_backup(CFG, engine, primary.assignment)
    path = write_personal_rosters(
        work / "personal.xlsx", CFG, engine, primary.assignment, backup.assignment
    )
    return engine, primary, backup, load_workbook(path)


def test_one_sheet_per_member(personal):
    engine, _, _, wb = personal
    assert wb.sheetnames == engine.members


def test_only_own_days_are_marked_and_no_other_names(personal):
    engine, primary, backup, wb = personal
    others = set(engine.members)
    for name in engine.members:
        ws = wb[name]
        marks = {}
        for row in ws.iter_rows(min_row=4, max_col=7):
            for cell in row:
                if cell.value in ("待機", "予備"):
                    marks[cell.row] = marks.get(cell.row, 0) + 1
                # 他者の氏名は一切出さない
                assert cell.value not in others - {name} or cell.value is None
        expected = sum(1 for d in engine.days if primary.assignment[d] == name) + sum(
            1 for d in engine.days if backup.assignment[d] == name
        )
        assert sum(marks.values()) == expected


def test_duty_is_red_and_backup_is_black(personal):
    engine, primary, backup, wb = personal
    for name in engine.members:
        ws = wb[name]
        for row in ws.iter_rows(min_row=4, max_col=7):
            for cell in row:
                if cell.value == "待機":
                    assert cell.font.color.rgb == "FFFF0000"
                elif cell.value == "予備":
                    assert cell.font.color.rgb == "FF000000"


def test_counts_are_total_duty_backup(personal):
    engine, primary, backup, wb = personal
    for name in engine.members:
        ws = wb[name]
        duty = sum(1 for d in engine.days if primary.assignment[d] == name)
        spare = sum(1 for d in engine.days if backup.assignment[d] == name)
        assert [ws.cell(row=3 + i, column=9).value for i in range(4)] == [
            "区分", "合計", "待機", "予備",
        ]
        assert [ws.cell(row=4 + i, column=10).value for i in range(3)] == [
            duty + spare, duty, spare,
        ]


def test_sheets_print_on_one_a4_landscape_page(personal):
    engine, _, _, wb = personal
    for name in engine.members:
        ws = wb[name]
        assert ws.page_setup.paperSize == 9
        assert ws.page_setup.orientation == "landscape"
        assert ws.page_setup.fitToWidth == 1
        assert ws.print_options.horizontalCentered
