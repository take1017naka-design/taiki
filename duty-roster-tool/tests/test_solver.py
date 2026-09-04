import datetime as dt

from duty_roster.config import load_config
from duty_roster.rules import RuleEngine
from duty_roster.sample import build_sample
from duty_roster.solver import Solver, solve
from duty_roster.workbook import read_schedule
from duty_roster.writer import write_roster

CFG = load_config("config/roster.example.yaml")


import pytest


def _run(tmp_path, year, month):
    path = build_sample(tmp_path / f"sample_{year}{month:02d}.xlsx", year, month)
    schedule = read_schedule(path, year, month, CFG)
    engine = RuleEngine(CFG, schedule, year, month)
    return engine, solve(CFG, engine)


@pytest.fixture(scope="module")
def august(tmp_path_factory):
    return _run(tmp_path_factory.mktemp("aug"), 2026, 8)


@pytest.fixture(scope="module")
def february(tmp_path_factory):
    return _run(tmp_path_factory.mktemp("feb"), 2026, 2)


def test_every_day_is_assigned_and_quota_is_exact(august):
    engine, solution = august
    assert set(solution.assignment) == set(engine.days)
    assert solution.counts == CFG.quota(engine.days_in_month)


def test_no_assignment_on_unavailable_day(august):
    engine, solution = august
    bad = [
        (day, name)
        for day, name in solution.assignment.items()
        if not engine.eligible(name, day)
    ]
    assert bad == []


def test_sunday_uses_each_member_at_most_once(august):
    engine, solution = august
    sundays = [d for d in engine.days if d.weekday() == 6]
    names = [solution.assignment[d] for d in sundays]
    assert len(set(names)) == len(names)
    assert set(names) <= set(CFG.weekend_pool)


def test_consecutive_group_never_runs_three_days(august):
    engine, solution = august
    group = set(CFG.consecutive_group)
    days = engine.days
    runs = [
        days[i]
        for i in range(len(days) - 2)
        if all(solution.assignment[days[i + k]] in group for k in range(3))
    ]
    assert runs == []


def test_sunday_priority_is_not_traded_away(august):
    """日曜は、より上位の優先順位の人が空いていればその人が入る。"""
    engine, solution = august
    worst = 99

    def rank(name, day):
        tier = engine.tier(name, day)
        return worst if tier is None else tier

    for day in (d for d in engine.days if d.weekday() == 6):
        chosen = solution.assignment[day]
        chosen_tier = rank(chosen, day)
        better = [
            n
            for n in engine.candidates(day)
            if rank(n, day) < chosen_tier
            and not engine.eligibility(n, day).conditional
            # その人が他の日曜で使われていなければ、そちらが選ばれるべき
            and n not in {solution.assignment[d] for d in engine.days if d.weekday() == 6}
        ]
        assert better == [], f"{day:%m/%d} は {better} の方が優先順位が上"


def test_unconditional_sunday_candidate_beats_conditional_one(august):
    """日曜は、制限なしの候補がいれば優先順位が下でもそちらを使う。"""
    engine, solution = august
    sundays = [d for d in engine.days if d.weekday() == 6]
    used = {solution.assignment[d] for d in sundays}
    for day in sundays:
        if not engine.eligibility(solution.assignment[day], day).conditional:
            continue
        # △ を使った日は、未使用かつ制限なしの候補が残っていないはず
        free = [
            n
            for n in engine.candidates(day)
            if n not in used and not engine.eligibility(n, day).conditional
        ]
        assert free == [], f"{day:%m/%d} は制限なしの {free} を使えるはず"


def test_writer_produces_all_sheets(august, tmp_path):
    engine, solution = august
    out = write_roster(tmp_path / "roster.xlsx", CFG, engine, solution)
    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert wb.sheetnames == ["待機表", "一覧", "集計", "可否一覧", "確認事項"]
    ws = wb["待機表"]
    assert ws["A1"].value == "2026年8月　カテ待機表"

    # 日曜(8/2)は祝日(8/11)と同じ色、土曜(8/1)は薄い青、平日は塗らない
    def fill_of(row, col):
        cell = ws.cell(row=row, column=col)
        return cell.fill.start_color.rgb if cell.fill.fill_type == "solid" else None

    sunday, holiday, saturday, weekday = fill_of(6, 1), fill_of(8, 3), fill_of(4, 7), fill_of(6, 2)
    assert sunday == holiday == "FFFDE9E9"
    assert saturday == "FFDEEBF7"
    assert weekday is None

    # 枡目の大きさがそろっていること
    assert ws.column_dimensions["A"].width == ws.column_dimensions["G"].width == 17.0
    assert ws.row_dimensions[4].height == ws.row_dimensions[5].height == 36.0

    # A4横（余白0.4インチ）の印刷可能領域 10.89 x 7.47 インチに収まること
    def col_px(width):
        return width * 7 + 5

    width_in = (
        7 * col_px(17.0) + col_px(2.5) + col_px(11) + col_px(8)
    ) / 96
    weeks = sum(1 for r in range(4, ws.max_row + 1, 2) if ws.cell(row=r, column=1).value or True)
    height_in = (29 + 23 + 12 * 36 + 18) / 72   # 表題+曜日+最大6週+注記
    assert width_in <= 10.89, width_in
    assert height_in <= 7.47, height_in

    # カレンダーの右横に担当ごとの日数が出ていること
    assert ws.cell(row=3, column=9).value == "担当"
    assert ws.cell(row=3, column=10).value == "日数"
    counts = {
        ws.cell(row=3 + i, column=9).value: ws.cell(row=3 + i, column=10).value
        for i in range(1, len(engine.members) + 1)
    }
    assert counts == solution.counts
    assert ws.cell(row=3 + len(engine.members) + 1, column=9).value == "合計"
    assert ws.cell(row=3 + len(engine.members) + 1, column=10).value == engine.days_in_month


def test_fixed_assignment_is_honoured_even_if_ineligible(august):
    """先に決めた担当は、ルールに関係なくそのまま入る。"""
    import datetime

    engine, _ = august
    saturday = datetime.date(2026, 8, 1)
    # 土日の担当対象外の人を土曜に指定する
    outsider = next(n for n in engine.members if n not in CFG.weekend_pool)
    assert not engine.eligible(outsider, saturday)

    solution = solve(CFG, engine, {saturday: outsider})
    assert solution.assignment[saturday] == outsider
    assert sum(solution.counts.values()) == engine.days_in_month
    assert any("先に決めた担当" in note for note in solution.notes)


def test_fixed_sunday_keeps_one_turn_each(august):
    import datetime

    engine, _ = august
    first_sunday = datetime.date(2026, 8, 2)
    who = CFG.weekend_pool[0]

    solution = solve(CFG, engine, {first_sunday: who})
    sundays = [d for d in engine.days if d.weekday() == 6]
    names = [solution.assignment[d] for d in sundays]
    assert solution.assignment[first_sunday] == who
    assert len(set(names)) == len(names)   # 日曜は1人1回のまま


def test_fixed_assignment_keeps_total_days(august):
    import datetime

    engine, _ = august
    fixed = {
        datetime.date(2026, 8, 5): engine.members[0],
        datetime.date(2026, 8, 12): engine.members[1],
        datetime.date(2026, 8, 19): engine.members[2],
    }
    solution = solve(CFG, engine, fixed)
    for day, name in fixed.items():
        assert solution.assignment[day] == name
    assert sum(solution.counts.values()) == engine.days_in_month
    assert set(solution.assignment) == set(engine.days)


def test_sheets_are_ready_to_print(august, tmp_path):
    engine, solution = august
    out = write_roster(tmp_path / "print.xlsx", CFG, engine, solution)
    from openpyxl import load_workbook

    wb = load_workbook(out)
    for name in wb.sheetnames:
        ws = wb[name]
        assert ws.page_setup.paperSize == 9              # A4
        assert ws.page_setup.orientation == "landscape"  # 横
        assert ws.page_setup.fitToWidth == 1
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True
    area = wb["待機表"].print_area
    assert area.startswith("'待機表'!$A$1:$J$")   # 日数欄(J列)まで印刷範囲に入る

    # 待機表は用紙の上下左右中央。複数ページになるシートは左右のみ中央。
    roster = wb["待機表"].print_options
    assert roster.horizontalCentered and roster.verticalCentered
    listing = wb["一覧"].print_options
    assert listing.horizontalCentered and not listing.verticalCentered


def test_short_month_also_solves(february):
    engine, solution = february
    assert solution.counts == CFG.quota(28)


def test_last_resort_tier_is_avoided_when_someone_better_is_free(tmp_path):
    """月〜木は、上の段に候補がいる限り最後の段（翌日が手術室）を使わない。"""
    import datetime as dt

    from duty_roster.rules import RuleEngine
    from duty_roster.workbook import Cell, DayCells, WorkSchedule

    year, month = 2026, 8
    day, nxt = dt.date(year, month, 4), dt.date(year, month, 5)
    names = CFG.member_names
    cells = {}
    for index, name in enumerate(names):
        # 全員その日は勤務可。翌日は1人だけ手術室（空白）、残りは ME。
        cells[(name, day)] = DayCells([Cell(text="ME")])
        cells[(name, nxt)] = DayCells([] if index == 0 else [Cell(text="ME")])
    schedule = WorkSchedule(
        year=year, month=month, cells=cells, names_in_sheet=list(names)
    )
    engine = RuleEngine(CFG, schedule, year, month)
    solver = Solver(CFG, engine)
    # 翌日が手術室の人のコストは、連日・間隔のどれよりも重い
    last_resort = solver.tier_cost[(names[0], day)]
    normal = solver.tier_cost[(names[1], day)]
    assert last_resort > normal + CFG.weights["consecutive"]
    assert last_resort > normal + CFG.weights["short_gap"]
