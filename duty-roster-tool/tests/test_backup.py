import datetime as dt

import pytest

from duty_roster.backup import solve_backup
from duty_roster.config import load_config
from duty_roster.rules import RuleEngine
from duty_roster.sample import build_sample
from duty_roster.solver import solve
from duty_roster.workbook import read_schedule
from duty_roster.writer import write_roster

CFG = load_config("config/roster.example.yaml")
YEAR, MONTH = 2026, 8


@pytest.fixture(scope="module")
def rosters(tmp_path_factory):
    path = build_sample(tmp_path_factory.mktemp("bk") / "sample.xlsx", YEAR, MONTH)
    engine = RuleEngine(CFG, read_schedule(path, YEAR, MONTH, CFG), YEAR, MONTH)
    primary = solve(CFG, engine)
    backup = solve_backup(CFG, engine, primary.assignment)
    return engine, primary, backup


def test_every_day_has_a_backup(rosters):
    engine, _, backup = rosters
    assert set(backup.assignment) == set(engine.days)


def test_backup_is_a_different_person(rosters):
    engine, primary, backup = rosters
    same = [d for d in engine.days if backup.assignment[d] == primary.assignment[d]]
    assert same == []


def test_anchor_covers_the_dependents(rosters):
    """待機が依存2名の日は、必ずバックアップ役が予備に入る。"""
    engine, primary, backup = rosters
    anchor = CFG.backup_anchor
    dependents = set(CFG.backup_dependents)
    days = [d for d in engine.days if primary.assignment[d] in dependents]
    assert days, "サンプルに依存2名の待機日がない"
    for day in days:
        assert backup.assignment[day] == anchor
        assert day in backup.forced


def test_forbidden_pairs_are_respected(rosters):
    engine, primary, backup = rosters
    for day in engine.days:
        blocked = CFG.backup_forbidden_pairs.get(primary.assignment[day], set())
        assert backup.assignment[day] not in blocked


def test_combined_runs_stay_within_the_limit(rosters):
    """待機と予備を合算した連続日数が上限を超えない。"""
    engine, primary, backup = rosters
    for name in engine.members:
        days = sorted(
            {d for d in engine.days if primary.assignment[d] == name}
            | {d for d in engine.days if backup.assignment[d] == name}
        )
        limit = CFG.backup_max_run(name)
        run = 1
        for earlier, later in zip(days, days[1:]):
            run = run + 1 if (later - earlier).days == 1 else 1
            assert run <= limit, f"{name} が {later:%m/%d} まで {run} 日連続（上限{limit}）"


def test_backup_avoids_unavailable_people(rosters):
    engine, _, backup = rosters
    for day in engine.days:
        name = backup.assignment[day]
        if engine.backup_eligible(name, day):
            continue
        # 候補が誰もいない日だけは埋めたうえで違反として報告する
        assert not [n for n in engine.members if engine.backup_eligible(n, day)]
        assert any(f"{day:%m/%d}" in v for v in backup.violations)


def test_backup_workbook_has_the_same_format(rosters, tmp_path):
    engine, primary, backup = rosters
    out = write_roster(
        tmp_path / "backup.xlsx",
        CFG,
        engine,
        backup,
        sheet_name="予備待機表",
        title="カテ予備待機表",
        partner=primary.assignment,
        availability_fn=engine.backup_eligibility,
    )
    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert wb.sheetnames == ["予備待機表", "一覧", "集計", "可否一覧", "確認事項"]
    ws = wb["予備待機表"]
    assert ws["A1"].value == f"{YEAR}年{MONTH}月　カテ予備待機表"
    assert ws.page_setup.orientation == "landscape"
    assert wb["一覧"]["D1"].value == "待機者"   # 待機者の列が入る


def test_next_day_operating_room_names_are_red(rosters, tmp_path):
    """予備が翌日手術室業務の担当者なら、氏名を赤字にする。"""
    engine, primary, backup = rosters
    out = write_roster(
        tmp_path / "red.xlsx",
        CFG,
        engine,
        backup,
        sheet_name="予備待機表",
        title="カテ予備待機表",
        partner=primary.assignment,
        name_color_fn=lambda day, name: (
            "FFFF0000" if engine.next_day_is_operating_room(name, day) else None
        ),
    )
    from openpyxl import load_workbook

    ws = load_workbook(out)["予備待機表"]
    checked = 0
    for row in range(4, ws.max_row + 1, 2):
        for col in range(1, 8):
            day_value = ws.cell(row=row, column=col).value
            name = ws.cell(row=row + 1, column=col).value
            if not isinstance(day_value, int) or not name:
                continue
            day = dt.date(YEAR, MONTH, day_value)
            font = ws.cell(row=row + 1, column=col).font
            red = bool(font.color and font.color.rgb == "FFFF0000")
            assert red == engine.next_day_is_operating_room(name, day)
            checked += 1
    assert checked == len(engine.days)


def test_history_records_and_resets(tmp_path):
    """日曜・祝日の予備の実績を月ごとに記録し、同じ月は置き換える。"""
    from duty_roster.history import History

    path = tmp_path / "history.json"
    history = History.load(path)

    def is_sunday(day):
        return day.weekday() == 6

    august = {dt.date(2026, 8, d): "担当B" for d in (2, 9)}
    history.record_holiday_backup(2026, 8, august, is_sunday)
    history.save()

    reloaded = History.load(path)
    assert reloaded.holiday_backup_totals() == {"担当B": 2}
    assert reloaded.recorded_months() == ["2026-08"]

    # 同じ月を作り直しても二重に数えない
    reloaded.record_holiday_backup(2026, 8, {dt.date(2026, 8, 2): "担当C"}, is_sunday)
    assert reloaded.holiday_backup_totals() == {"担当C": 1}

    # 対象月を除いた集計（その月を組み直すときに使う）
    reloaded.record_holiday_backup(2026, 9, {dt.date(2026, 9, 6): "担当D"}, is_sunday)
    assert reloaded.holiday_backup_totals(exclude_month=(2026, 9)) == {"担当C": 1}


def test_holiday_backup_uses_the_history(rosters):
    """通算回数が少ない人が、日曜・祝日の予備に選ばれやすくなる。"""
    engine, primary, _ = rosters
    holidays = [d for d in engine.days if engine.is_red_day(d)]
    assert holidays

    loaded = {n: 10 for n in engine.members}
    light = CFG.weekend_pool[0]
    loaded[light] = 0

    biased = solve_backup(CFG, engine, primary.assignment, None, loaded)
    assert light in [biased.assignment[d] for d in holidays]


def test_weekend_excluded_never_appears_on_saturday_or_sunday(rosters):
    """土日は予備の対象外にした人が入らない（祝日と重なる日を除く）。"""
    engine, _, backup = rosters
    excluded = CFG.backup_weekend_excluded
    assert excluded, "サンプル設定に weekend_excluded がない"
    for day in engine.days:
        if day.weekday() not in (5, 6) or engine.is_holiday(day):
            continue
        assert backup.assignment[day] not in excluded, f"{day:%m/%d} に {backup.assignment[day]}"


def test_holiday_assignment_of_consult_member_is_reported():
    """祝日に「要相談」の人を予備にしたら確認事項に出る。"""
    from duty_roster.backup import BackupSolver
    from duty_roster.rules import RuleEngine
    from duty_roster.workbook import DayCells, WorkSchedule

    holiday = dt.date(2026, 8, 11)   # 山の日
    schedule = WorkSchedule(2026, 8, {}, CFG.member_names)
    engine = RuleEngine(CFG, schedule, 2026, 8)
    consult = sorted(CFG.backup_holiday_consult)[0]
    other = next(n for n in CFG.member_names if n != consult)

    primary = {d: other for d in engine.days}
    solver = BackupSolver(CFG, engine, primary)
    assignment = {d: consult for d in engine.days}
    messages = solver.collect_violations(assignment)
    assert any(f"{holiday:%m/%d}" in m and "相談" in m for m in messages)


def test_prefer_more_takes_the_larger_share(rosters):
    """均等割りが割り切れないとき、指定した人は多いほうの日数にする。"""
    engine = rosters[0]
    cfg = load_config("config/roster.example.yaml")
    target = cfg.member_names[1]
    cfg.raw["backup_roster"]["even_quota_prefer_more"] = [target]
    primary = solve(cfg, engine)
    backup = solve_backup(cfg, engine, primary.assignment)

    pool = [n for n in engine.members if n not in cfg.backup_quota_ignore]
    counts = {n: backup.counts.get(n, 0) for n in pool}
    total = sum(counts.values())
    if total % len(pool):
        assert counts[target] == -(-total // len(pool))
    else:
        assert counts[target] == total // len(pool)


def test_history_can_be_imported_from_a_backup_roster(rosters, tmp_path):
    """作成済みの予備待機表から、その月の日曜・祝日の実績を取り込める。"""
    from duty_roster.history import History

    engine, primary, backup = rosters
    path = write_roster(
        tmp_path / "予備待機表.xlsx",
        CFG,
        engine,
        backup,
        sheet_name="予備待機表",
        title="カテ予備待機表",
        partner=primary.assignment,
    )
    history = History.load(tmp_path / "history.json")
    year, month, counts = history.import_backup_roster(path)

    assert (year, month) == (YEAR, MONTH)
    expected: dict[str, int] = {}
    for day in engine.days:
        if engine.is_red_day(day):
            name = backup.assignment[day]
            expected[name] = expected.get(name, 0) + 1
    assert counts == expected
    assert history.recorded_months() == [f"{YEAR}-{MONTH:02d}"]
